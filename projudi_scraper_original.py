
import pandas as pd
import csv
import threading
import queue
from datetime import datetime
import os
import re
import io
import socket
import subprocess
import shutil
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaIoBaseDownload, MediaFileUpload
from seleniumbase import SB
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from fake_useragent import UserAgent
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import time
import logging
import traceback
import tempfile
import service_account_loader  # Importar o módulo para carregar credenciais
import zipfile

# Escopos necessários para acessar o Google Drive
SCOPES = ['https://www.googleapis.com/auth/drive']

# Configuração global do user agent
DEFAULT_USER_AGENT = "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36"
try:
    ua = UserAgent()
    user_agent = ua.random
except Exception:
    ua = None
    user_agent = DEFAULT_USER_AGENT

PROJUDI_LOGIN_URL = "https://projudi.tjgo.jus.br/LogOn?PaginaAtual=-200"
PROJUDI_BUSCA_URL = "https://projudi.tjgo.jus.br/BuscaProcesso?PaginaAtual=4"
PROJUDI_PROCESSO_INPUT_XPATH = '/html/body/div/form/div/fieldset/div[4]/input'


class DriverRestartNeeded(Exception):
    """Sinaliza que o driver do Selenium precisa ser reinicializado."""
    pass


class BrowserDegradationDetected(Exception):
    """Sinaliza que o navegador está degradado (muitas falhas consecutivas) e precisa ser reiniciado."""
    pass

def porta_debug_livre():
    """Retorna uma porta TCP livre para o Chrome DevTools usar no worker."""
    with socket.socket(socket.AF_INET, socket.SOCK_STREAM) as s:
        s.bind(("", 0))
        return s.getsockname()[1]


def limpar_processos_chrome_orfaos(logger=None, aguardar_liberacao=True):
    """Limpa TODOS os processos Chrome/Chromium e aguarda até que sejam realmente encerrados."""
    try:
        # Tenta várias vezes até garantir que todos os processos foram mortos
        for tentativa in range(3):
            # Mata TODOS os processos chrome
            subprocess.run(["pkill", "-9", "chrome"], capture_output=True, timeout=10)
            subprocess.run(["pkill", "-9", "chromium"], capture_output=True, timeout=10)
            subprocess.run(["pkill", "-9", "chromedriver"], capture_output=True, timeout=10)
            subprocess.run(["killall", "-9", "chrome"], capture_output=True, timeout=10)
            subprocess.run(["killall", "-9", "chromium"], capture_output=True, timeout=10)
            subprocess.run(["killall", "-9", "chromedriver"], capture_output=True, timeout=10)

            # Verifica se ainda existem processos chrome rodando
            result = subprocess.run(["pgrep", "-f", "chrome"], capture_output=True, timeout=10)
            if result.returncode != 0:  # Nenhum processo encontrado
                break

            if logger:
                logger.info(f"Ainda existem processos Chrome. Tentativa {tentativa + 1}/3 de limpeza...")
            time.sleep(2)

        if aguardar_liberacao:
            # Aguarda tempo suficiente para o sistema liberar recursos (file descriptors, memória, etc.)
            time.sleep(5)

        if logger:
            logger.info("Limpeza de processos Chrome concluída. Sistema pronto para reiniciar.")
    except Exception as e:
        if logger:
            logger.debug(f"Erro ao limpar processos Chrome: {e}")


def configurar_driver(sb):
    """Configura o driver com as configurações necessárias"""
    # Desabilita esperas implícitas para máxima velocidade
    sb.driver.implicitly_wait(0)
    
    # Configura page load strategy para não esperar carregamento completo
    sb.driver.execute_cdp_cmd('Page.enable', {})
    sb.driver.execute_cdp_cmd('Network.enable', {})
    
    # Define o user agent
    sb.driver.execute_cdp_cmd('Network.setUserAgentOverride', {
        "userAgent": user_agent
    })

    # Desabilita a detecção de DevTools aberto para evitar que o Chrome feche
    sb.driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
        'source': '''
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            window.chrome = {
                runtime: {}
            };
        '''
    })

class Application:
    def __init__(self):
        self.excel_dir = None
        self.porcentagem_concluida = 0
        self.processo_atual = "Inicializando..."
        self.salvando = ""
        self.processing_thread = None
        self.stop_processing = False
        self.total_registros = 0
        self.lock = threading.Lock()
        self.projudi_usuario = os.environ.get("PROJUDI_USUARIO", "07228313151")
        self.projudi_senha = os.environ.get("PROJUDI_SENHA", "Senhaprojudi24.")
        self.thread_local = threading.local()
        self.max_navegadores = 4
        self.setup_logging()
        self.ultimo_salvamento_drive = 0
        self.intervalo_salvamento_drive = 100  # Salvar no Drive a cada 100 processos
        self.arquivo_drive_id = None  # Para armazenar o ID do arquivo no Drive
        self.processos_nao_encontrados = []
        self.processos_nao_encontrados_set = set()
        self.processos_concluidos = set()
        self.processos_falha_motivos = {}
        # Upload assíncrono para o Drive
        self.drive_upload_queue = queue.Queue()
        self.drive_upload_thread = None
        self.drive_upload_running = False
        # Controle de sessões ativas do navegador para evitar limpeza global concorrente.
        self.sessoes_browser_ativas = 0
        self.sessoes_browser_lock = threading.Lock()
        self.chrome_cleanup_lock = threading.Lock()

    def setup_logging(self):
        self.logger = logging.getLogger('Application')
        self.logger.setLevel(logging.DEBUG)

        if self.logger.handlers:
            return

        # Configurar log para arquivo no diretório do projeto
        file_handler = None
        log_path = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'application.log')
        try:
            file_handler = logging.FileHandler(log_path, mode='w')
            file_handler.setLevel(logging.DEBUG)
        except Exception:
            # Fallback para /tmp quando o diretório atual não for gravável
            try:
                tmp_log_path = os.path.join(tempfile.gettempdir(), 'application.log')
                file_handler = logging.FileHandler(tmp_log_path, mode='w')
                file_handler.setLevel(logging.DEBUG)
            except Exception:
                file_handler = None

        # Configurar log para console
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)

        # Definir formato do log
        formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        if file_handler:
            file_handler.setFormatter(formatter)
            self.logger.addHandler(file_handler)
        console_handler.setFormatter(formatter)
        self.logger.addHandler(console_handler)

    def _get_pendentes_path(self):
        base_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), ".context")
        try:
            os.makedirs(base_dir, exist_ok=True)
            return os.path.join(base_dir, "processos_pendentes.csv")
        except Exception:
            return os.path.join(os.path.dirname(os.path.abspath(__file__)), "processos_pendentes.csv")

    def _marcar_sessao_browser_ativa(self, contexto, worker_id):
        with self.sessoes_browser_lock:
            self.sessoes_browser_ativas += 1
            total = self.sessoes_browser_ativas
        self.logger.debug(f"[{contexto} {worker_id}] Sessões Chrome ativas: {total}")

    def _desmarcar_sessao_browser_ativa(self, contexto, worker_id):
        with self.sessoes_browser_lock:
            if self.sessoes_browser_ativas > 0:
                self.sessoes_browser_ativas -= 1
            total = self.sessoes_browser_ativas
        self.logger.debug(f"[{contexto} {worker_id}] Sessões Chrome ativas após encerramento: {total}")

    def _limpar_processos_chrome_com_seguranca(self, contexto, worker_id):
        """Evita matar sessões ativas de outros workers ao limpar processos órfãos."""
        with self.chrome_cleanup_lock:
            with self.sessoes_browser_lock:
                sessoes_ativas = self.sessoes_browser_ativas

            if sessoes_ativas > 0:
                self.logger.warning(
                    f"[{contexto} {worker_id}] Limpeza global do Chrome ignorada: "
                    f"há {sessoes_ativas} sessão(ões) ativa(s)."
                )
                return False

            self.logger.info(f"[{contexto} {worker_id}] Limpando processos Chrome órfãos (sem sessões ativas).")
            limpar_processos_chrome_orfaos(self.logger)
            return True

    def _registrar_motivo_falha(self, numero_linha, numero_processo, motivo):
        if not motivo:
            return
        chave = (numero_linha, numero_processo)
        with self.lock:
            self.processos_falha_motivos[chave] = motivo

    def registrar_concluido(self, numero_linha, numero_processo):
        chave = (numero_linha, numero_processo)
        with self.lock:
            self.processos_concluidos.add(chave)
            if chave in self.processos_falha_motivos:
                del self.processos_falha_motivos[chave]
            if chave in self.processos_nao_encontrados_set:
                self.processos_nao_encontrados_set.discard(chave)

    def registrar_nao_encontrado(self, numero_linha, numero_processo, motivo=None):
        chave = (numero_linha, numero_processo)
        with self.lock:
            if chave in self.processos_concluidos:
                return
            if chave not in self.processos_nao_encontrados_set:
                self.processos_nao_encontrados.append((numero_linha, numero_processo))
                self.processos_nao_encontrados_set.add(chave)
            if motivo:
                self.processos_falha_motivos[chave] = motivo
        if motivo:
            self.logger.warning(
                f"Processo {numero_processo} (linha {numero_linha + 1}) não foi localizado e será reprocessado. Motivo: {motivo}"
            )
        else:
            self.logger.warning(
                f"Processo {numero_processo} (linha {numero_linha + 1}) não foi localizado e será reprocessado."
            )

    def _adicionar_pendentes_faltantes(self, processos_com_indice):
        with self.lock:
            concluidos = set(self.processos_concluidos)
            nao_encontrados = set(self.processos_nao_encontrados_set)
        pendentes = [p for p in processos_com_indice if p not in concluidos and p not in nao_encontrados]
        if pendentes:
            self.logger.warning(
                f"Foram encontrados {len(pendentes)} processos pendentes sem registro de falha. "
                f"Adicionando ao reprocessamento."
            )
            for numero_linha, numero_processo in pendentes:
                self.registrar_nao_encontrado(numero_linha, numero_processo, motivo="pendente_sem_registro")

    def exportar_processos_pendentes(self):
        if not self.processos_nao_encontrados:
            return
        caminho = self._get_pendentes_path()
        try:
            with open(caminho, "w", newline="") as arquivo:
                writer = csv.writer(arquivo)
                writer.writerow(["linha", "numero_processo", "motivo"])
                for numero_linha, numero_processo in self.processos_nao_encontrados:
                    motivo = self.processos_falha_motivos.get((numero_linha, numero_processo), "")
                    writer.writerow([numero_linha + 1, numero_processo, motivo])
            self.logger.warning(f"Lista de processos pendentes salva em: {caminho}")
        except Exception as e:
            self.logger.error(f"Erro ao salvar lista de pendentes: {e}")

    def iniciar_upload_assincrono(self, service):
        """Inicia a thread de upload assíncrono para o Drive."""
        self.drive_upload_running = True
        self.drive_upload_thread = threading.Thread(
            target=self._worker_upload_drive,
            args=(service,),
            daemon=True
        )
        self.drive_upload_thread.start()
        self.logger.info("Thread de upload assíncrono iniciada.")

    def _worker_upload_drive(self, service):
        """Worker que processa uploads para o Drive em background."""
        while self.drive_upload_running or not self.drive_upload_queue.empty():
            try:
                # Espera por item na fila com timeout para verificar stop
                try:
                    file_path = self.drive_upload_queue.get(timeout=2)
                except queue.Empty:
                    continue

                if file_path is None:  # Sinal para encerrar
                    break

                self._executar_upload_drive(service, file_path)
                self.drive_upload_queue.task_done()

            except Exception as e:
                self.logger.error(f"Erro no worker de upload: {e}")

    def _executar_upload_drive(self, service, file_path):
        """Executa o upload real para o Drive."""
        try:
            self.logger.info("Iniciando upload assíncrono para o Google Drive...")
            media = MediaFileUpload(file_path, resumable=True)
            service.files().update(fileId=self.arquivo_drive_id, media_body=media).execute()
            self.logger.info("Upload assíncrono concluído com sucesso.")
        except Exception as e:
            self.logger.error(f"Erro no upload assíncrono: {str(e)}")

    def agendar_upload_drive(self, file_path):
        """Agenda um upload para ser executado assincronamente."""
        self.drive_upload_queue.put(file_path)

    def parar_upload_assincrono(self):
        """Para a thread de upload e aguarda finalizar uploads pendentes."""
        self.drive_upload_running = False
        self.drive_upload_queue.put(None)  # Sinal para encerrar
        if self.drive_upload_thread and self.drive_upload_thread.is_alive():
            self.drive_upload_thread.join(timeout=60)
            self.logger.info("Thread de upload assíncrono encerrada.")

    def start_processing(self):
        # Carregar credenciais do service account
        if not service_account_loader.load_service_account():
            self.logger.error("Não foi possível carregar as credenciais do service_account.json")
            return
        
        print("Olá, estamos iniciando a busca dos processos")
        self.logger.info("Olá, estamos iniciando a busca dos processos")
        self.logger.info("Iniciando processamento...")
        self.processing_thread = threading.Thread(target=self.processar_planilha)
        self.processing_thread.start()
        self.processing_thread.join()  # Espera a thread terminar
        self.logger.info("Processamento concluído. Encerrando o programa.")

    def authenticate_google_drive(self):
        env_path = os.environ.get("GOOGLE_APPLICATION_CREDENTIALS")
        if env_path and os.path.exists(env_path):
            caminho_service_account = env_path
        else:
            caminho_atual = os.path.dirname(os.path.abspath(__file__))
            caminho_service_account = os.path.join(caminho_atual, 'service_account.json')

        if not os.path.exists(caminho_service_account):
            raise FileNotFoundError(f"service_account.json não encontrado em: {caminho_service_account}")

        # Use 'caminho_service_account' onde você normalmente usaria o nome do arquivo
        creds = Credentials.from_service_account_file(caminho_service_account, scopes=SCOPES)
        return creds

    def get_latest_excel_file(self, service):
        folder_id = '1DXQxsOW4MVJIn_kEWlbya7zx6ICUXRd0'
        query = f"'{folder_id}' in parents and mimeType='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'"

        results = service.files().list(q=query, spaces='drive',
                                       fields='files(id, name, createdTime)',
                                       orderBy='createdTime desc').execute()
        items = results.get('files', [])
        if not items:
            raise Exception('Nenhuma planilha encontrada no Google Drive.')

        arquivo = items[0]
        file_id = arquivo['id']
        file_name = arquivo['name']

        request = service.files().get_media(fileId=file_id)
        fh = io.FileIO(file_name, 'wb')
        downloader = MediaIoBaseDownload(fh, request)
        done = False
        while not done:
            status, done = downloader.next_chunk()
            self.logger.info(f"Download {int(status.progress() * 100)}%.")

        return file_name

    def baixar_planilha_valida(self, service, max_tentativas=3):
        """Baixa a planilha do Drive garantindo que não veio corrompida."""
        for tentativa in range(1, max_tentativas + 1):
            caminho = self.get_latest_excel_file(service)
            try:
                wb = load_workbook(caminho, read_only=True)
                wb.close()
                return caminho
            except (zipfile.BadZipFile, Exception) as e:
                self.logger.error(f"Planilha baixada está corrompida (tentativa {tentativa}/{max_tentativas}): {e}")
                self.limpar_arquivo_local(caminho)
                if tentativa == max_tentativas:
                    raise
                time.sleep(1)

    def upload_file_to_google_drive(self, service, file_path):
        file_metadata = {
            'name': os.path.basename(file_path),
            'parents': ['1DXQxsOW4MVJIn_kEWlbya7zx6ICUXRd0']
        }
        media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        self.logger.info(f"Arquivo '{file_path}' carregado com sucesso no Google Drive. ID: {file.get('id')}")

    def limpar_arquivo_local(self, arquivo):
        try:
            if os.path.exists(arquivo):
                os.remove(arquivo)
                self.logger.info(f"Arquivo local '{arquivo}' removido com sucesso")
        except Exception as e:
            self.logger.error(f"Erro ao remover arquivo local '{arquivo}': {str(e)}")



    def _set_login_realizado(self, valor: bool):
        setattr(self.thread_local, 'login_realizado', valor)

    def _reset_login_realizado(self):
        self._set_login_realizado(False)

    def _login_ja_realizado(self):
        return getattr(self.thread_local, 'login_realizado', False)



    def realizar_login_projudi(self, driver):
        if self._login_ja_realizado():
            return

        max_tentativas_login = 3
        for tentativa in range(max_tentativas_login):
            try:
                self.logger.info(f"Realizando login no PROJUDI... (tentativa {tentativa + 1}/{max_tentativas_login})")
                try:
                    driver.get(PROJUDI_LOGIN_URL)
                except Exception as nav_error:
                    nav_msg = f"{nav_error.__class__.__name__}: {str(nav_error)}"
                    if self._is_driver_connection_error(nav_msg):
                        raise DriverRestartNeeded(nav_msg)
                    pass  # Continua mesmo se timeout

                # Tenta encontrar elementos imediatamente
                login_input = None
                for tentativa_login in range(5):
                    try:
                        login_input = driver.find_element(By.XPATH, '//*[@id="login"]')
                        break
                    except NoSuchElementException:
                        if tentativa_login < 4:
                            time.sleep(0.05)
                if not login_input:
                    login_input = self.esperar_condicao(driver, EC.presence_of_element_located((By.XPATH, '//*[@id="login"]')), timeout=1.5, tentativas=3)
                login_input.clear()
                login_input.send_keys(self.projudi_usuario)

                senha_input = None
                try:
                    senha_input = driver.find_element(By.XPATH, '//*[@id="senha"]')
                except NoSuchElementException:
                    senha_input = self.esperar_condicao(driver, EC.presence_of_element_located((By.XPATH, '//*[@id="senha"]')), timeout=1.5, tentativas=3)
                senha_input.clear()
                senha_input.send_keys(self.projudi_senha)

                entrar_button = None
                try:
                    entrar_button = driver.find_element(By.XPATH, '//*[@id="formLogin"]/div[4]/input[1]')
                except NoSuchElementException:
                    entrar_button = self.esperar_condicao(driver, EC.element_to_be_clickable((By.XPATH, '//*[@id="formLogin"]/div[4]/input[1]')), timeout=1.5, tentativas=3)
                entrar_button.click()

                # Aguarda mudança de URL rapidamente
                for tentativa_url in range(10):
                    if "LogOn" not in driver.current_url:
                        break
                    time.sleep(0.1)

                self._selecionar_perfil_usuario(driver)
                self._fechar_popups_projudi(driver)

                try:
                    driver.get(PROJUDI_BUSCA_URL)
                except Exception as nav_error:
                    nav_msg = f"{nav_error.__class__.__name__}: {str(nav_error)}"
                    if self._is_driver_connection_error(nav_msg):
                        raise DriverRestartNeeded(nav_msg)
                    pass  # Continua mesmo se timeout
                self._set_login_realizado(True)
                self.logger.info("Login no PROJUDI concluído e página de busca carregada.")
                return  # Sucesso, sai da função
            except DriverRestartNeeded:
                raise
            except Exception as e:
                error_message = str(e)
                self.logger.error(f"Falha ao realizar login no PROJUDI (tentativa {tentativa + 1}/{max_tentativas_login}): {error_message}")
                driver_error_msg = f"{e.__class__.__name__}: {error_message}"
                if self._is_driver_connection_error(driver_error_msg):
                    raise DriverRestartNeeded(driver_error_msg)
                if tentativa < max_tentativas_login - 1:
                    self.logger.info(f"Aguardando 5s antes de tentar login novamente...")
                    time.sleep(5)
                else:
                    self.logger.debug(traceback.format_exc())
                    raise

    def _selecionar_perfil_usuario(self, driver):
        try:
            opcoes_xpath = [
                "//fieldset//a[contains(@href, 'SelecionarPerfil')]",
                "/html/body/div[3]/fieldset/label[1]/a",
                "/html/body/div[3]/fieldset/label[2]/a"
            ]

            for xpath in opcoes_xpath:
                try:
                    link = self.esperar_condicao(driver, EC.element_to_be_clickable((By.XPATH, xpath)), timeout=1.2, tentativas=4)
                    link.click()
                    self.logger.info("Perfil selecionado no PROJUDI.")
                    return
                except TimeoutException:
                    continue
        except Exception as e:
            self.logger.warning(f"Não foi possível selecionar o perfil automaticamente: {str(e)}")

    def _fechar_popups_projudi(self, driver):
        possiveis_popups = [
            "/html/body/div[2]/div[3]/div/button",
            "//button[contains(@class, 'close')]"
        ]

        for xpath in possiveis_popups:
            try:
                elemento = driver.find_element(By.XPATH, xpath)
                if elemento.is_displayed():
                    driver.execute_script("arguments[0].click();", elemento)
                    self.logger.debug(f"Pop-up fechado via xpath: {xpath}")
            except NoSuchElementException:
                continue
            except Exception as e:
                self.logger.debug(f"Falha ao fechar pop-up ({xpath}): {str(e)}")

    def _elemento_presente(self, driver, by, locator):
        try:
            driver.find_element(by, locator)
            return True
        except NoSuchElementException:
            return False

    def _is_driver_connection_error(self, message):
        if not message:
            return False
        lower_msg = message.lower()
        driver_error_keywords = [
            'failed to establish a new connection',
            'connection refused',
            'invalid session id',
            'invalidsessionidexception',
            'chrome not reachable',
            'disconnected: not connected to devtools',
            'cannot connect to chrome',
            'httpconnectionpool',
            'target closed',
            'session deleted because of page crash',
            'webview not found',
            'no such window'
        ]
        return any(keyword in lower_msg for keyword in driver_error_keywords)

    def esperar_condicao(self, driver, condition, timeout=0.3, tentativas=2):
        ultimo_erro = None
        for _ in range(tentativas):
            try:
                return WebDriverWait(driver, timeout).until(condition)
            except TimeoutException as exc:
                ultimo_erro = exc
        if ultimo_erro:
            raise ultimo_erro

    def alternar_para_formulario_busca(self, driver):
        driver.switch_to.default_content()
        if self._elemento_presente(driver, By.XPATH, PROJUDI_PROCESSO_INPUT_XPATH):
            return

        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for iframe in iframes:
            driver.switch_to.default_content()
            driver.switch_to.frame(iframe)
            if self._elemento_presente(driver, By.XPATH, PROJUDI_PROCESSO_INPUT_XPATH):
                self.logger.debug("Formulário de busca localizado dentro de iframe.")
                return

        driver.switch_to.default_content()

    def alternar_para_detalhes_processo(self, driver):
        driver.switch_to.default_content()
        detalhes_xpaths = [
            '//fieldset[contains(@id, "Autor") or contains(., "Autor")]',
            '/html/body/div[2]/form/div[1]/fieldset/fieldset/fieldset[1]/fieldset',
            '/html/body/div[4]/form/div[1]/fieldset/fieldset/fieldset[1]/fieldset',
            '/html/body/div[3]/form/div[1]/fieldset/fieldset/fieldset[1]/fieldset'
        ]

        for xpath in detalhes_xpaths:
            if self._elemento_presente(driver, By.XPATH, xpath):
                return

        iframes = driver.find_elements(By.TAG_NAME, "iframe")
        for iframe in iframes:
            driver.switch_to.default_content()
            driver.switch_to.frame(iframe)
            for xpath in detalhes_xpaths:
                if self._elemento_presente(driver, By.XPATH, xpath):
                    self.logger.debug("Detalhes do processo localizados dentro de iframe.")
                    return

        driver.switch_to.default_content()


    def processar_planilha(self):
        try:
            creds = self.authenticate_google_drive()
            service = build('drive', 'v3', credentials=creds)

            self.excel_dir = self.baixar_planilha_valida(service)

            self.logger.info("Lendo planilha Excel...")
            df_original = pd.read_excel(self.excel_dir, engine='openpyxl')
            self.logger.info(f"Planilha lida com sucesso. Total de registros: {len(df_original)}")

            data_atual = datetime.now().strftime("%d de %B %Y")
            novo_nome_arquivo = f"TJ GO - último andamento ({data_atual}).xlsx"

            novo_caminho_arquivo = novo_nome_arquivo

            self.logger.info("Preparando DataFrame...")
            df_novo = pd.DataFrame()
            df_novo['Numero_Processo'] = df_original.iloc[:, 0]

            if 'Advogados' in df_original.columns:
                df_novo['Advogados'] = df_original['Advogados']
            else:
                df_novo['Advogados'] = df_original.iloc[:, 1]

            novos_campos = ['Autor', 'Reu', 'Data_Distribuicao', 'Valor_Causa', 'Assunto',
                            'Comarca', 'Data_Recente', 'Andamento_Recente', 'Usuario_Ultima_Movimentacao']
            for campo in novos_campos:
                df_novo[campo] = ""

            self.total_registros = len(df_novo)
            processos = df_novo['Numero_Processo'].tolist()
            self.logger.info(f"DataFrame preparado. Total de processos para buscar: {self.total_registros}")

            # Criar o arquivo no Drive e obter o ID
            self.logger.info("Criando arquivo no Google Drive...")
            self.arquivo_drive_id = self.criar_arquivo_drive(service, novo_nome_arquivo)
            self.logger.info(f"Arquivo criado no Drive com ID: {self.arquivo_drive_id}")

            # Iniciar thread de upload assíncrono
            self.iniciar_upload_assincrono(service)

            print("Abrindo navegador Chrome... Isso pode levar alguns segundos.")
            processos_com_indice = list(enumerate(processos))

            fila_processos = queue.Queue()
            for processo in processos_com_indice:
                fila_processos.put(processo)

            total_workers = min(self.max_navegadores, fila_processos.qsize())
            self.logger.info(f"Iniciando processamento paralelo com {total_workers} navegadores de automação (fila compartilhada).")

            threads = []
            for worker_id in range(1, total_workers + 1):
                thread = threading.Thread(
                    target=self.processar_lote,
                    args=(worker_id, fila_processos, df_novo, service, novo_caminho_arquivo),
                    daemon=True
                )
                threads.append(thread)
                thread.start()
                # Delay entre iniciar cada worker para não sobrecarregar o sistema
                if worker_id < total_workers:
                    self.logger.info(f"Aguardando 10s antes de iniciar próximo worker...")
                    time.sleep(10)

            for thread in threads:
                thread.join()

            # Identificar processos perdidos e tentar novamente os não encontrados
            self._adicionar_pendentes_faltantes(processos_com_indice)
            self.reprocessar_nao_encontrados(df_novo)
            if self.processos_nao_encontrados:
                self.exportar_processos_pendentes()

            if not self.stop_processing:
                self.salvar_planilha(df_novo, novo_caminho_arquivo)
                self.ajustar_largura_colunas(novo_caminho_arquivo)
                self.comparar_e_destacar(df_original, df_novo, novo_caminho_arquivo)
                self.logger.info(f"Acompanhamento de processos concluído com sucesso!\nNova planilha salva como: {novo_nome_arquivo}")

                # Aguarda uploads pendentes e faz upload final
                self.parar_upload_assincrono()
                self.salvar_no_drive(service, novo_caminho_arquivo)

                # Remove os arquivos locais após salvar no Drive
                self.limpar_arquivo_local(self.excel_dir)  # Remove o arquivo original
                self.limpar_arquivo_local(novo_caminho_arquivo)  # Remove o novo arquivo

        except Exception as e:
            self.logger.error(f"Erro crítico ao processar a planilha: {str(e)}")
            self.logger.debug(traceback.format_exc())

        finally:
            self.logger.info("Processamento finalizado.")
            # Encerra o programa após o processamento
            os._exit(0)

    def dividir_processos_em_lotes(self, processos_com_indice, num_lotes):
        if not processos_com_indice:
            return []
        quantidade_lotes = max(1, min(num_lotes, len(processos_com_indice)))
        lotes = [[] for _ in range(quantidade_lotes)]
        for posicao, processo in enumerate(processos_com_indice):
            lotes[posicao % quantidade_lotes].append(processo)
        return [lote for lote in lotes if lote]

    def processar_lote(self, worker_id, fila_processos, df_novo, service, caminho_arquivo):
        tentativas_falha_consecutivas = 0
        max_tentativas_falha = 10  # Máximo de falhas consecutivas antes de desistir
        backoff_base = 2  # Segundos base para backoff exponencial

        # Parâmetros de detecção de degradação
        DEGRADACAO_LIMITE_FALHAS = 8  # Falhas consecutivas na mesma sessão para considerar degradação
        DEGRADACAO_DELAY_REINICIO = 15  # Segundos de espera ao reiniciar por degradação

        PAUSA_RECUPERACAO_BASE = 120  # Segundos base de pausa quando atinge muitas falhas

        while not self.stop_processing:
            # Verificar se atingiu máximo de falhas consecutivas
            if tentativas_falha_consecutivas >= max_tentativas_falha:
                # Limpa processos Chrome travados antes de tentar novamente
                self._limpar_processos_chrome_com_seguranca("WORKER", worker_id)

                # Pausa escalonada: cada worker espera um tempo diferente para não voltarem juntos
                pausa_escalonada = PAUSA_RECUPERACAO_BASE + (worker_id * 30)
                self.logger.warning(
                    f"[WORKER {worker_id}] Atingiu {max_tentativas_falha} falhas consecutivas. "
                    f"Fazendo pausa de {pausa_escalonada}s antes de tentar novamente..."
                )
                time.sleep(pausa_escalonada)
                tentativas_falha_consecutivas = 0  # Reseta o contador após a pausa
                continue

            try:
                self._reset_login_realizado()

                # Limpa processos Chrome travados ANTES de cada tentativa
                if tentativas_falha_consecutivas > 0:
                    self._limpar_processos_chrome_com_seguranca("WORKER", worker_id)
                    # Delay escalonado por worker para não abrirem todos ao mesmo tempo
                    delay_escalonado = worker_id * 5
                    self.logger.info(f"[WORKER {worker_id}] Aguardando {delay_escalonado}s antes de reabrir...")
                    time.sleep(delay_escalonado)

                # Gera novo user agent a cada nova sessão do navegador
                if ua:
                    try:
                        novo_user_agent = ua.random
                    except Exception as ua_error:
                        self.logger.warning(f"[WORKER {worker_id}] Falha ao gerar user agent: {ua_error}. Usando fallback.")
                        novo_user_agent = user_agent
                else:
                    novo_user_agent = user_agent
                self.logger.info(f"[WORKER {worker_id}] Abrindo navegador com novo user agent.")

                debug_port = porta_debug_livre()
                user_data_dir = tempfile.mkdtemp(prefix=f"sb_worker_{worker_id}_")
                chromium_args = [
                    f"--remote-debugging-port={debug_port}",
                    "--disable-dev-shm-usage",
                    "--no-sandbox",
                    f"--user-data-dir={user_data_dir}",
                ]
                try:
                    with SB(
                        uc=True,
                        incognito=True,
                        headless=True,
                        # Evita conflito de porta do DevTools entre múltiplos workers
                        chromium_arg=",".join(chromium_args),
                    ) as sb:
                        self._marcar_sessao_browser_ativa("WORKER", worker_id)
                        try:
                            # Configura driver com user agent específico desta sessão
                            sb.driver.implicitly_wait(0)
                            sb.driver.execute_cdp_cmd('Page.enable', {})
                            sb.driver.execute_cdp_cmd('Network.enable', {})
                            sb.driver.execute_cdp_cmd('Network.setUserAgentOverride', {"userAgent": novo_user_agent})
                            sb.driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
                                'source': '''
                                    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                                    window.chrome = {runtime: {}};
                                '''
                            })

                            sb.driver.set_page_load_timeout(5)
                            self.realizar_login_projudi(sb.driver)
                            wait = WebDriverWait(sb.driver, 2)

                            # Resetar contador de falhas após sucesso na abertura do navegador
                            tentativas_falha_consecutivas = 0

                            # Contador de falhas DENTRO desta sessão do navegador
                            falhas_sessao = 0
                            processos_sucesso_sessao = 0

                            while not self.stop_processing:
                                try:
                                    numero_linha, numero_processo = fila_processos.get_nowait()
                                except queue.Empty:
                                    return

                                self.logger.info(f"[WORKER {worker_id}] Linha {numero_linha + 1}/{self.total_registros} - Processo: {numero_processo}")

                                try:
                                    dados = self.extrair_dados_processo(sb.driver, wait, numero_processo)
                                except DriverRestartNeeded as driver_exc:
                                    self.logger.warning(
                                        f"[WORKER {worker_id}] Driver reiniciado durante o processo {numero_processo}. Devolvendo para fila. Detalhes: {driver_exc}"
                                    )
                                    fila_processos.put((numero_linha, numero_processo))
                                    raise driver_exc
                                except Exception as e:
                                    error_message = str(e)
                                    self.logger.error(f"[WORKER {worker_id}] Erro ao processar o processo {numero_processo}: {error_message}")
                                    self.logger.debug(traceback.format_exc())
                                    motivo = f"erro_processamento: {error_message}" if error_message else "erro_processamento"
                                    if len(motivo) > 200:
                                        motivo = motivo[:200]
                                    self.registrar_nao_encontrado(numero_linha, numero_processo, motivo=motivo)
                                    falhas_sessao += 1

                                    # Verifica degradação: muitas falhas consecutivas na mesma sessão
                                    if falhas_sessao >= DEGRADACAO_LIMITE_FALHAS:
                                        self.logger.warning(
                                            f"[WORKER {worker_id}] DEGRADAÇÃO DETECTADA: {falhas_sessao} falhas consecutivas na sessão. "
                                            f"Reiniciando navegador com novo user agent..."
                                        )
                                        raise BrowserDegradationDetected(f"{falhas_sessao} falhas consecutivas")
                                    continue

                                if not dados or 'Erro' in dados:
                                    motivo = "erro_extracao"
                                    if isinstance(dados, dict) and dados.get('Erro'):
                                        motivo = f"erro_extracao: {dados.get('Erro')}"
                                    if len(motivo) > 200:
                                        motivo = motivo[:200]
                                    self.registrar_nao_encontrado(numero_linha, numero_processo, motivo=motivo)
                                    falhas_sessao += 1

                                    # Verifica degradação
                                    if falhas_sessao >= DEGRADACAO_LIMITE_FALHAS:
                                        self.logger.warning(
                                            f"[WORKER {worker_id}] DEGRADAÇÃO DETECTADA: {falhas_sessao} falhas consecutivas na sessão. "
                                            f"Reiniciando navegador com novo user agent..."
                                        )
                                        raise BrowserDegradationDetected(f"{falhas_sessao} falhas consecutivas")
                                else:
                                    # Sucesso! Resetar contador de falhas da sessão
                                    if falhas_sessao > 0:
                                        self.logger.info(f"[WORKER {worker_id}] Recuperado após {falhas_sessao} falhas. Contador resetado.")
                                    falhas_sessao = 0
                                    processos_sucesso_sessao += 1
                                    self.atualizar_resultados(df_novo, numero_linha, numero_processo, dados, service, caminho_arquivo)
                        finally:
                            self._desmarcar_sessao_browser_ativa("WORKER", worker_id)
                finally:
                    shutil.rmtree(user_data_dir, ignore_errors=True)

            except BrowserDegradationDetected as degradacao:
                # Degradação detectada - aguarda mais tempo e reinicia com novo perfil
                self.logger.warning(
                    f"[WORKER {worker_id}] Navegador degradado ({degradacao}). "
                    f"Aguardando {DEGRADACAO_DELAY_REINICIO}s para evitar detecção anti-bot..."
                )
                time.sleep(DEGRADACAO_DELAY_REINICIO)
                # Não incrementa tentativas_falha_consecutivas pois é reinício preventivo
                continue
            except DriverRestartNeeded as driver_error:
                tentativas_falha_consecutivas += 1
                delay = min(backoff_base * (2 ** tentativas_falha_consecutivas), 60)  # Max 60 segundos
                self.logger.warning(
                    f"[WORKER {worker_id}] Driver do Selenium foi encerrado. Aguardando {delay}s antes de reabrir (tentativa {tentativas_falha_consecutivas}/{max_tentativas_falha}). ({driver_error})"
                )
                time.sleep(delay)
                continue
            except Exception as worker_error:
                tentativas_falha_consecutivas += 1
                delay = min(backoff_base * (2 ** tentativas_falha_consecutivas), 60)  # Max 60 segundos
                self.logger.error(f"[WORKER {worker_id}] Falha inesperada (tentativa {tentativas_falha_consecutivas}/{max_tentativas_falha}). Aguardando {delay}s antes de reabrir. Detalhes: {worker_error}")
                self.logger.debug(traceback.format_exc())
                time.sleep(delay)
                continue

    def atualizar_resultados(self, df, numero_linha, numero_processo, dados, service, caminho_arquivo):
        with self.lock:
            self.preencher_dados_no_df(df, numero_linha, dados)
            if self.total_registros > 0:
                self.porcentagem_concluida += 1 / self.total_registros
            self.processo_atual = numero_processo

            progresso_atual = int(self.porcentagem_concluida * self.total_registros)
            if progresso_atual and progresso_atual % 50 == 0:
                self.logger.info(f"Progresso: {int(self.porcentagem_concluida * 100)}% concluído")
                self.salvar_planilha(df, caminho_arquivo)
                self.ajustar_largura_colunas(caminho_arquivo)
                self.salvando = f"Salvando... Última vez salvo: {datetime.now().strftime('%H:%M')}"

            if progresso_atual - self.ultimo_salvamento_drive >= self.intervalo_salvamento_drive:
                # Upload assíncrono - não bloqueia os workers
                self.agendar_upload_drive(caminho_arquivo)
                self.ultimo_salvamento_drive = progresso_atual
        self.registrar_concluido(numero_linha, numero_processo)

    def extrair_elemento_com_timeout(self, wait, xpath, timeout=0.1):
        try:
            # Tenta primeiro com o xpath fornecido
            try:
                return WebDriverWait(wait._driver, timeout).until(
                    EC.presence_of_element_located((By.XPATH, xpath))
                ).text.strip()
            except:
                # Se falhar e o xpath contiver div[4], tenta com div[3]
                if 'div[4]' in xpath:
                    xpath_alternativo = xpath.replace('div[4]', 'div[3]')
                    return WebDriverWait(wait._driver, timeout).until(
                        EC.presence_of_element_located((By.XPATH, xpath_alternativo))
                    ).text.strip()
                # Se falhar e o xpath contiver div[3], tenta com div[4]
                elif 'div[3]' in xpath:
                    xpath_alternativo = xpath.replace('div[3]', 'div[4]')
                    return WebDriverWait(wait._driver, timeout).until(
                        EC.presence_of_element_located((By.XPATH, xpath_alternativo))
                    ).text.strip()
                else:
                    raise
        except TimeoutException:
            return "Não encontrado (timeout)"
        except Exception as e:
            error_message = str(e)
            if self._is_driver_connection_error(error_message):
                raise DriverRestartNeeded(error_message)
            return f"Erro: {error_message}"

    def obter_texto_primeiro_xpath(self, wait, xpaths, timeout=0.5):
        for xpath in xpaths:
            valor = self.extrair_elemento_com_timeout(wait, xpath, timeout)
            if valor and not valor.startswith("Não encontrado") and not valor.startswith("Erro"):
                return valor
        return ""

    def extrair_dados_processo(self, driver, wait, numero_processo):
        tentativa = 0
        max_tentativas = 5
        self.logger.info(f"[INICIANDO] Extração de dados para o processo: {numero_processo}")

        while tentativa < max_tentativas:
            try:
                # Otimização: só navega se não estiver já na página de busca
                current_url = driver.current_url
                ja_na_pagina_busca = "BuscaProcesso" in current_url or "PaginaAtual=4" in current_url

                # Se não é a primeira tentativa ou não está na página de busca, navega
                if tentativa > 0 or not ja_na_pagina_busca:
                    self.logger.info(f"[NAVEGANDO] Para a página de busca do TJ-GO...")
                    if tentativa > 0:
                        self.logger.info(f"[REFRESH] Reacessando página de busca (tentativa {tentativa + 1})...")
                    try:
                        driver.get(PROJUDI_BUSCA_URL)
                    except Exception:
                        # Mesmo se timeout, continua tentando encontrar elementos
                        pass
                driver.switch_to.default_content()
                
                self.logger.info(f"[LOCALIZANDO] Campo de número do processo...")
                # Tenta encontrar imediatamente sem espera - múltiplas tentativas rápidas
                processo_input = None
                for tentativa_campo in range(8):
                    try:
                        processo_input = driver.find_element(By.XPATH, PROJUDI_PROCESSO_INPUT_XPATH)
                        break
                    except NoSuchElementException:
                        try:
                            processo_input = driver.find_element(By.XPATH, '//*[@id="ProcessoNumero"]')
                            break
                        except NoSuchElementException:
                            # Verifica iframes rapidamente
                            try:
                                iframes = driver.find_elements(By.TAG_NAME, "iframe")
                                for iframe in iframes:
                                    driver.switch_to.default_content()
                                    driver.switch_to.frame(iframe)
                                    try:
                                        processo_input = driver.find_element(By.XPATH, PROJUDI_PROCESSO_INPUT_XPATH)
                                        break
                                    except:
                                        driver.switch_to.default_content()
                                        continue
                                if processo_input:
                                    break
                            except:
                                pass
                    if processo_input:
                        break
                    if tentativa_campo < 7:
                        time.sleep(0.03)  # Espera mínima entre tentativas
                
                if not processo_input:
                    # Última tentativa com espera
                    processo_input = self.esperar_condicao(
                        driver,
                        EC.presence_of_element_located((By.XPATH, PROJUDI_PROCESSO_INPUT_XPATH)),
                        tentativas=2
                    )
                processo_input.clear()

                self.logger.info(f"[DIGITANDO] Número do processo: {numero_processo}")
                processo_input.send_keys(numero_processo)

                # Clicar no botão para limpar filtro de arquivados (sempre mostrar arquivados)
                try:
                    filter_button = None
                    try:
                        filter_button = driver.find_element(By.XPATH, "/html/body/div/form/div/fieldset/div[5]/fieldset/div[1]/label/button[2]")
                    except NoSuchElementException:
                        try:
                            filter_button = driver.find_element(By.CSS_SELECTOR, "button[name='imaLimparProcessoStatus']")
                        except NoSuchElementException:
                            filter_button = self.esperar_condicao(driver, EC.element_to_be_clickable((By.XPATH, "/html/body/div/form/div/fieldset/div[5]/fieldset/div[1]/label/button[2]")), timeout=0.3, tentativas=2)
                    if filter_button:
                        driver.execute_script("arguments[0].click();", filter_button)
                        self.logger.debug("Botão de limpar status de arquivados clicado.")
                except Exception as status_error:
                    self.logger.debug(f"Não foi possível clicar no botão de arquivados: {status_error}")

                self.logger.info(f"[CLICANDO] Botão de busca...")
                # Tenta clicar no botão em uma das duas posições possíveis
                botao_clicado = False
                erros_botao = []

                # Lista de estratégias para localizar o botão
                estrategias_botao = [
                    ('XPATH', '//input[@name="imgSubmeter"]'),  # Por name - mais confiável
                    ('XPATH', '//input[@value="Buscar"]'),  # Por value
                    ('CSS', '#divBotoesCentralizados > input[type=submit]:nth-child(1)'),  # CSS específico
                    ('CSS', 'input[name="imgSubmeter"]'),  # CSS por name
                    ('XPATH', '/html/body/div[3]/form/div/fieldset/div[5]/input[1]'),
                    ('XPATH', '/html/body/div[3]/form/div/fieldset/div[5]/input[2]'),
                    ('XPATH', '//input[@type="submit" and contains(@value, "Pesquisar")]'),
                    ('XPATH', '//input[@type="submit" and contains(@value, "Buscar")]'),
                    ('CSS', 'input[type="submit"][value*="Buscar"]'),
                    ('XPATH', '//input[@type="submit"]')
                ]

                # Salva a URL antes do clique
                url_antes_clique = driver.current_url

                for tipo, localizador in estrategias_botao:
                    try:
                        botao_busca = None
                        # Tenta encontrar imediatamente primeiro
                        try:
                            if tipo == 'XPATH':
                                botao_busca = driver.find_element(By.XPATH, localizador)
                            else:
                                botao_busca = driver.find_element(By.CSS_SELECTOR, localizador)
                        except NoSuchElementException:
                            # Só então tenta com espera
                            if tipo == 'XPATH':
                                botao_busca = self.esperar_condicao(driver, EC.element_to_be_clickable((By.XPATH, localizador)), timeout=0.3, tentativas=2)
                            else:
                                botao_busca = self.esperar_condicao(driver, EC.element_to_be_clickable((By.CSS_SELECTOR, localizador)), timeout=0.3, tentativas=2)

                        # Tenta primeiro clique normal
                        clique_sucesso = False
                        try:
                            botao_busca.click()
                            clique_sucesso = True
                        except Exception as click_err:
                            # Se falhar, tenta JavaScript click
                            try:
                                driver.execute_script("arguments[0].click();", botao_busca)
                                clique_sucesso = True
                            except:
                                pass

                        if not clique_sucesso:
                            # Última tentativa: executa o JavaScript do onclick diretamente
                            try:
                                driver.execute_script("AlterarValue('PaginaAtual','2'); VerificarCampos();")
                                clique_sucesso = True
                            except:
                                pass

                        if clique_sucesso:
                            botao_clicado = True
                            self.logger.info(f"[SUCESSO] Botão clicado")
                            break
                        else:
                            # Se não conseguiu clicar de jeito nenhum, continua para próxima estratégia
                            continue
                    except Exception as e:
                        erro_msg = str(e)
                        erros_botao.append(f"{tipo} - {localizador}: {erro_msg[:80]}")
                        self.logger.debug(f"[DEBUG] Botão não encontrado por {tipo} {localizador}: {erro_msg[:100]}")
                        continue

                if not botao_clicado:
                    self.logger.error("[ERRO] Não encontrou o botão de busca")
                    raise Exception(f"Não foi possível encontrar o botão de busca")

                # Tenta localizar o elemento de dados do processo em diferentes posições
                elemento_encontrado = False
                timeout = 0.3
                erros_dados = []

                self.alternar_para_detalhes_processo(driver)

                for xpath_dados in ['/html/body/div[2]/form/div[1]/fieldset/fieldset/fieldset[1]/fieldset',
                                   '/html/body/div[4]/form/div[1]/fieldset/fieldset/fieldset[1]/fieldset',
                                   '/html/body/div[3]/form/div[1]/fieldset/fieldset/fieldset[1]/fieldset',
                                   '//fieldset[contains(@id, "Autor") or contains(., "Autor")]',
                                   '//fieldset//fieldset//fieldset[1]']:
                    try:
                        self.esperar_condicao(
                            driver,
                            EC.presence_of_element_located((By.XPATH, xpath_dados)),
                            timeout=timeout,
                            tentativas=2
                        )
                        elemento_encontrado = True
                        break
                    except:
                        continue

                if not elemento_encontrado:
                    self.logger.error("[ERRO] Dados do processo não encontrados")
                    raise Exception(f"Dados do processo não encontrados")

                dados = {}

                campos_xpaths = {
                    'Autor': [
                        '/html/body/div[2]/form/div[1]/fieldset/fieldset/fieldset[1]/fieldset/span[1]',
                        '/html/body/div[4]/form/div[1]/fieldset/fieldset/fieldset[1]/fieldset',
                        '/html/body/div[3]/form/div[1]/fieldset/fieldset/fieldset[1]/fieldset'
                    ],
                    'Reu': [
                        '/html/body/div[2]/form/div[1]/fieldset/fieldset/fieldset[2]/fieldset/span[1]',
                        '/html/body/div[4]/form/div[1]/fieldset/fieldset/fieldset[2]/fieldset',
                        '/html/body/div[3]/form/div[1]/fieldset/fieldset/fieldset[2]/fieldset'
                    ],
                    'Comarca': [
                        '/html/body/div[2]/form/div[1]/fieldset/fieldset/fieldset[3]/span[1]',
                        '/html/body/div[4]/form/div[1]/fieldset/fieldset/fieldset[3]/span[1]'
                    ],
                    'Valor_Causa': [
                        '/html/body/div[2]/form/div[1]/fieldset/fieldset/fieldset[3]/span[4]',
                        '/html/body/div[4]/form/div[1]/fieldset/fieldset/fieldset[3]/span[4]'
                    ],
                    'Assunto': [
                        '/html/body/div[2]/form/div[1]/fieldset/fieldset/fieldset[3]/span[3]/table/tbody/tr/td',
                        '/html/body/div[4]/form/div[1]/fieldset/fieldset/fieldset[3]/span[3]/table/tbody/tr/td'
                    ],
                    'Data_Distribuicao': [
                        '/html/body/div[2]/form/div[1]/fieldset/fieldset/fieldset[3]/span[8]',
                        '/html/body/div[4]/form/div[1]/fieldset/fieldset/fieldset[3]/span[8]'
                    ],
                    'Data_Recente': [
                        '/html/body/div[2]/form/div[1]/div/div[1]/table/tbody/tr[1]/td[3]',
                        '/html/body/div[4]/form/div[1]/div/div[1]/table/tbody/tr[1]/td[3]'
                    ],
                    'Andamento_Recente': [
                        '/html/body/div[2]/form/div[1]/div/div[1]/table/tbody/tr[1]/td[2]',
                        '/html/body/div[4]/form/div[1]/div/div[1]/table/tbody/tr[1]/td[2]'
                    ],
                    'Usuario_Ultima_Movimentacao': [
                        '/html/body/div[2]/form/div[1]/div/div[1]/table/tbody/tr[1]/td[4]',
                        '/html/body/div[2]/form/div[1]/div/div[1]/table/tbody/tr[1]/td[3]',
                        '/html/body/div[4]/form/div[1]/div/div[1]/table/tbody/tr[1]/td[4]'
                    ]
                }

                for campo, xpaths in campos_xpaths.items():
                    dados[campo] = self.obter_texto_primeiro_xpath(wait, xpaths, timeout=0.1)

                if dados['Autor']:
                    dados['Autor'] = re.sub(r'Raça:.*', '', dados['Autor'], flags=re.IGNORECASE).strip()
                    unwanted_words = ['Nome', 'Social']
                    for word in unwanted_words:
                        dados['Autor'] = dados['Autor'].replace(word, '')
                    dados['Autor'] = dados['Autor'].strip()
                    names = [name.strip() for name in dados['Autor'].split('\n') if name.strip()]
                    unique_names = list(dict.fromkeys(names))
                    dados['Autor'] = '\n'.join(unique_names)

                if dados['Reu']:
                    unwanted_words = ['Nome', 'Raça', 'Social']
                    for word in unwanted_words:
                        dados['Reu'] = dados['Reu'].replace(word, '')
                    dados['Reu'] = dados['Reu'].strip()
                    names = [name.strip() for name in dados['Reu'].split('\n') if name.strip()]
                    unique_names = list(dict.fromkeys(names))
                    dados['Reu'] = '\n'.join(unique_names)

                if dados['Data_Distribuicao']:
                    date_str = dados['Data_Distribuicao'].split()[0]
                    try:
                        dados['Data_Distribuicao'] = pd.to_datetime(date_str, dayfirst=True).strftime('%d/%m/%Y')
                    except:
                        dados['Data_Distribuicao'] = ""

                if dados['Data_Recente']:
                    date_str = dados['Data_Recente'].split()[0]
                    try:
                        dados['Data_Recente'] = pd.to_datetime(date_str, dayfirst=True).strftime('%d/%m/%Y')
                    except:
                        dados['Data_Recente'] = ""

                for campo in ['Comarca', 'Valor_Causa', 'Assunto', 'Andamento_Recente', 'Usuario_Ultima_Movimentacao']:
                    if isinstance(dados[campo], str):
                        dados[campo] = dados[campo].strip()

                self.logger.info(f"[CONCLUIDO] Extração bem-sucedida para o processo {numero_processo}")
                return dados
            except DriverRestartNeeded as driver_error:
                raise driver_error
            except Exception as e:
                error_message = str(e)
                driver_error_msg = f"{e.__class__.__name__}: {error_message}"
                if self._is_driver_connection_error(driver_error_msg):
                    raise DriverRestartNeeded(driver_error_msg)
                tentativa += 1

                # Log mais detalhado sobre o erro
                if "ProcessoNumero" in error_message:
                    self.logger.warning(f"[ERRO] Tentativa {tentativa}: Não conseguiu localizar o campo de processo (possível Cloudflare)")
                elif "botão" in error_message.lower() or "input" in error_message.lower():
                    self.logger.warning(f"[ERRO] Tentativa {tentativa}: Não conseguiu localizar o botão de busca")
                else:
                    self.logger.warning(f"[ERRO] Tentativa {tentativa} falhou para o processo {numero_processo}: {error_message}")

                if tentativa == max_tentativas:
                    self.logger.error(f"Falha ao extrair dados após {max_tentativas} tentativas para o processo {numero_processo}")
                    return {'Erro': f"Falha ao extrair dados após {max_tentativas} tentativas: {error_message}"}



    def salvar_planilha(self, df, nome_arquivo):
        df.to_excel(nome_arquivo, index=False, engine='openpyxl')

    def ajustar_largura_colunas(self, nome_arquivo):
        workbook = load_workbook(nome_arquivo)
        sheet = workbook.active

        for column in sheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column_letter].width = adjusted_width

        workbook.save(nome_arquivo)

    def comparar_e_destacar(self, df_original, df_novo, nome_arquivo):
        workbook = load_workbook(nome_arquivo)
        sheet = workbook.active

        for date_col in ['Data_Distribuicao', 'Data_Recente']:
            if date_col in df_original.columns:
                df_original[date_col] = pd.to_datetime(df_original[date_col], errors='coerce').dt.date
            if date_col in df_novo.columns:
                df_novo[date_col] = pd.to_datetime(df_novo[date_col], errors='coerce').dt.date

        green_fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

        colunas_comparar = ['Autor', 'Reu', 'Data_Distribuicao', 'Valor_Causa', 'Assunto',
                            'Comarca', 'Data_Recente', 'Andamento_Recente', 'Usuario_Ultima_Movimentacao']

        for col, coluna in enumerate(colunas_comparar, start=3):
            if coluna in df_original.columns and coluna in df_novo.columns:
                for row in range(len(df_novo)):
                    valor_original = df_original.at[row, coluna]
                    valor_novo = df_novo.at[row, coluna]

                    cell = sheet.cell(row=row+2, column=col)

                    if pd.isna(valor_original) and not pd.isna(valor_novo):
                        cell.fill = green_fill
                    elif pd.isna(valor_novo) and not pd.isna(valor_original):
                        cell.fill = green_fill
                    elif pd.notna(valor_original) and pd.notna(valor_novo):
                        if coluna in ['Data_Distribuicao', 'Data_Recente']:
                            if valor_original != valor_novo:
                                cell.fill = green_fill
                            else:
                                cell.fill = PatternFill(fill_type=None)
                        else:
                            if str(valor_original).strip() != str(valor_novo).strip():
                                cell.fill = green_fill
                            else:
                                cell.fill = PatternFill(fill_type=None)
                    else:
                        cell.fill = PatternFill(fill_type=None)
            else:
                for row in range(2, sheet.max_row + 1):
                    cell = sheet.cell(row=row, column=col)
                    cell.fill = green_fill

        workbook.save(nome_arquivo)

    def criar_arquivo_drive(self, service, nome_arquivo):
        file_metadata = {
            'name': nome_arquivo,
            'parents': ['1DXQxsOW4MVJIn_kEWlbya7zx6ICUXRd0'],
            'mimeType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        }
        file = service.files().create(body=file_metadata, fields='id').execute()
        return file.get('id')

    def salvar_no_drive(self, service, file_path):
        try:
            self.logger.info("Iniciando salvamento no Google Drive...")
            media = MediaFileUpload(file_path, resumable=True)
            service.files().update(fileId=self.arquivo_drive_id, media_body=media).execute()
            self.logger.info("Arquivo atualizado com sucesso no Google Drive.")
        except Exception as e:
            self.logger.error(f"Erro ao salvar no Google Drive: {str(e)}")
            self.logger.debug(traceback.format_exc())

    def reprocessar_nao_encontrados(self, df_novo):
        """Reprocessa processos não encontrados usando múltiplos workers em paralelo."""
        if not self.processos_nao_encontrados:
            return

        with self.lock:
            self.processos_nao_encontrados = list(dict.fromkeys(self.processos_nao_encontrados))
            concluidos = set(self.processos_concluidos)
        self.processos_nao_encontrados = [p for p in self.processos_nao_encontrados if p not in concluidos]
        self.processos_nao_encontrados_set = set(self.processos_nao_encontrados)
        if not self.processos_nao_encontrados:
            return

        self.logger.info(f"Tentando reprocessar {len(self.processos_nao_encontrados)} processos não encontrados em paralelo.")
        max_tentativas = 3

        for tentativa in range(max_tentativas):
            if not self.processos_nao_encontrados:
                break

            self.logger.info(f"Reprocessamento - Tentativa {tentativa + 1}/{max_tentativas}")

            # Criar fila compartilhada com processos pendentes
            fila_reprocessamento = queue.Queue()
            for processo in self.processos_nao_encontrados:
                fila_reprocessamento.put(processo)

            # Lista thread-safe para processos ainda não encontrados
            processos_ainda_nao_encontrados = []
            lock_nao_encontrados = threading.Lock()

            # Determinar número de workers (máximo 2 para reprocessamento, pois são menos itens)
            num_workers = min(2, len(self.processos_nao_encontrados), self.max_navegadores)

            threads = []
            for worker_id in range(1, num_workers + 1):
                thread = threading.Thread(
                    target=self._worker_reprocessamento,
                    args=(worker_id, fila_reprocessamento, df_novo, processos_ainda_nao_encontrados, lock_nao_encontrados, tentativa + 1),
                    daemon=True
                )
                threads.append(thread)
                thread.start()
                # Delay entre iniciar cada worker para não sobrecarregar o sistema
                if worker_id < num_workers:
                    self.logger.info(f"[REPROCESSAMENTO] Aguardando 10s antes de iniciar próximo worker...")
                    time.sleep(10)

            # Aguardar todas as threads terminarem
            for thread in threads:
                thread.join()

            processos_unicos = list(dict.fromkeys(processos_ainda_nao_encontrados))
            with self.lock:
                self.processos_nao_encontrados = processos_unicos
                self.processos_nao_encontrados_set = set(processos_unicos)

            if not self.processos_nao_encontrados:
                self.logger.info("Todos os processos foram reprocessados com sucesso!")
                break

        if self.processos_nao_encontrados:
            self.logger.warning(f"Não foi possível processar {len(self.processos_nao_encontrados)} processos após {max_tentativas} tentativas.")

    def _worker_reprocessamento(self, worker_id, fila, df_novo, lista_nao_encontrados, lock_lista, tentativa_atual):
        """Worker para reprocessamento paralelo."""
        tentativas_falha_consecutivas = 0
        max_tentativas_falha = 5
        backoff_base = 2

        # Parâmetros de detecção de degradação
        DEGRADACAO_LIMITE_FALHAS = 5  # Menos tolerante no reprocessamento
        DEGRADACAO_DELAY_REINICIO = 10
        PAUSA_RECUPERACAO_BASE = 120  # Segundos base de pausa quando atinge muitas falhas

        while True:
            # Verificar se atingiu máximo de falhas consecutivas
            if tentativas_falha_consecutivas >= max_tentativas_falha:
                # Limpa processos Chrome travados antes de tentar novamente
                self._limpar_processos_chrome_com_seguranca("REPROCESSAMENTO WORKER", worker_id)

                # Pausa escalonada: cada worker espera um tempo diferente para não voltarem juntos
                pausa_escalonada = PAUSA_RECUPERACAO_BASE + (worker_id * 30)
                self.logger.warning(
                    f"[REPROCESSAMENTO WORKER {worker_id}] Atingiu {max_tentativas_falha} falhas consecutivas. "
                    f"Fazendo pausa de {pausa_escalonada}s antes de tentar novamente..."
                )
                time.sleep(pausa_escalonada)
                tentativas_falha_consecutivas = 0  # Reseta o contador após a pausa
                continue

            try:
                self._reset_login_realizado()

                # Limpa processos Chrome travados ANTES de cada tentativa
                if tentativas_falha_consecutivas > 0:
                    self._limpar_processos_chrome_com_seguranca("REPROCESSAMENTO WORKER", worker_id)
                    # Delay escalonado por worker para não abrirem todos ao mesmo tempo
                    delay_escalonado = worker_id * 5
                    self.logger.info(f"[REPROCESSAMENTO WORKER {worker_id}] Aguardando {delay_escalonado}s antes de reabrir...")
                    time.sleep(delay_escalonado)

                # Gera novo user agent a cada nova sessão
                if ua:
                    try:
                        novo_user_agent = ua.random
                    except Exception as ua_error:
                        self.logger.warning(f"[REPROCESSAMENTO WORKER {worker_id}] Falha ao gerar user agent: {ua_error}. Usando fallback.")
                        novo_user_agent = user_agent
                else:
                    novo_user_agent = user_agent
                self.logger.info(f"[REPROCESSAMENTO WORKER {worker_id}] Abrindo navegador com novo user agent...")

                debug_port = porta_debug_livre()
                user_data_dir = tempfile.mkdtemp(prefix=f"sb_reprocess_{worker_id}_")
                chromium_args = [
                    f"--remote-debugging-port={debug_port}",
                    "--disable-dev-shm-usage",
                    "--no-sandbox",
                    f"--user-data-dir={user_data_dir}",
                ]
                try:
                    with SB(
                        uc=True,
                        incognito=True,
                        headless=True,
                        chromium_arg=",".join(chromium_args),
                    ) as sb:
                        self._marcar_sessao_browser_ativa("REPROCESSAMENTO WORKER", worker_id)
                        try:
                            # Configura driver com user agent específico desta sessão
                            sb.driver.implicitly_wait(0)
                            sb.driver.execute_cdp_cmd('Page.enable', {})
                            sb.driver.execute_cdp_cmd('Network.enable', {})
                            sb.driver.execute_cdp_cmd('Network.setUserAgentOverride', {"userAgent": novo_user_agent})
                            sb.driver.execute_cdp_cmd('Page.addScriptToEvaluateOnNewDocument', {
                                'source': '''
                                    Object.defineProperty(navigator, 'webdriver', {get: () => undefined});
                                    window.chrome = {runtime: {}};
                                '''
                            })

                            sb.driver.set_page_load_timeout(5)
                            self.realizar_login_projudi(sb.driver)
                            wait = WebDriverWait(sb.driver, 2)

                            # Resetar contador após sucesso
                            tentativas_falha_consecutivas = 0

                            # Contador de falhas dentro desta sessão
                            falhas_sessao = 0

                            while True:
                                try:
                                    numero_linha, numero_processo = fila.get_nowait()
                                except queue.Empty:
                                    return  # Fila vazia, encerrar worker

                                self.logger.info(f"[REPROCESSAMENTO WORKER {worker_id}] Reprocessando: {numero_processo}")

                                try:
                                    dados = self.extrair_dados_processo(sb.driver, wait, numero_processo)
                                except DriverRestartNeeded as driver_exc:
                                    self.logger.warning(
                                        f"[REPROCESSAMENTO WORKER {worker_id}] Driver reiniciado. Devolvendo processo para fila."
                                    )
                                    fila.put((numero_linha, numero_processo))
                                    raise driver_exc
                                except Exception as e:
                                    self.logger.error(f"[REPROCESSAMENTO WORKER {worker_id}] Erro ao reprocessar {numero_processo}: {e}")
                                    with lock_lista:
                                        lista_nao_encontrados.append((numero_linha, numero_processo))
                                    self._registrar_motivo_falha(numero_linha, numero_processo, f"erro_reprocessamento: {e}")
                                    falhas_sessao += 1

                                    # Verifica degradação
                                    if falhas_sessao >= DEGRADACAO_LIMITE_FALHAS:
                                        self.logger.warning(
                                            f"[REPROCESSAMENTO WORKER {worker_id}] DEGRADAÇÃO DETECTADA: {falhas_sessao} falhas consecutivas. "
                                            f"Reiniciando navegador..."
                                        )
                                        raise BrowserDegradationDetected(f"{falhas_sessao} falhas consecutivas")
                                    continue

                                if dados and 'Erro' not in dados:
                                    with self.lock:
                                        self.preencher_dados_no_df(df_novo, numero_linha, dados)
                                    self.registrar_concluido(numero_linha, numero_processo)
                                    self.logger.info(f"[REPROCESSAMENTO WORKER {worker_id}] Processo {numero_processo} reprocessado com sucesso!")
                                    # Sucesso - reseta contador
                                    if falhas_sessao > 0:
                                        self.logger.info(f"[REPROCESSAMENTO WORKER {worker_id}] Recuperado após {falhas_sessao} falhas.")
                                    falhas_sessao = 0
                                else:
                                    with lock_lista:
                                        lista_nao_encontrados.append((numero_linha, numero_processo))
                                    self._registrar_motivo_falha(numero_linha, numero_processo, "erro_reprocessamento")
                                    falhas_sessao += 1

                                    # Verifica degradação
                                    if falhas_sessao >= DEGRADACAO_LIMITE_FALHAS:
                                        self.logger.warning(
                                            f"[REPROCESSAMENTO WORKER {worker_id}] DEGRADAÇÃO DETECTADA: {falhas_sessao} falhas consecutivas. "
                                            f"Reiniciando navegador..."
                                        )
                                        raise BrowserDegradationDetected(f"{falhas_sessao} falhas consecutivas")
                        finally:
                            self._desmarcar_sessao_browser_ativa("REPROCESSAMENTO WORKER", worker_id)
                finally:
                    shutil.rmtree(user_data_dir, ignore_errors=True)

            except BrowserDegradationDetected as degradacao:
                # Degradação detectada - aguarda e reinicia com novo perfil
                self.logger.warning(
                    f"[REPROCESSAMENTO WORKER {worker_id}] Navegador degradado ({degradacao}). "
                    f"Aguardando {DEGRADACAO_DELAY_REINICIO}s..."
                )
                time.sleep(DEGRADACAO_DELAY_REINICIO)
                continue
            except DriverRestartNeeded:
                tentativas_falha_consecutivas += 1
                delay = min(backoff_base * (2 ** tentativas_falha_consecutivas), 30)
                self.logger.warning(
                    f"[REPROCESSAMENTO WORKER {worker_id}] Driver reiniciado. Aguardando {delay}s (tentativa {tentativas_falha_consecutivas}/{max_tentativas_falha})"
                )
                time.sleep(delay)
                continue
            except Exception as e:
                tentativas_falha_consecutivas += 1
                delay = min(backoff_base * (2 ** tentativas_falha_consecutivas), 30)
                self.logger.error(f"[REPROCESSAMENTO WORKER {worker_id}] Falha inesperada. Aguardando {delay}s. Detalhes: {e}")
                time.sleep(delay)
                continue

    def preencher_dados_no_df(self, df, numero_linha, dados):
        for campo, valor in dados.items():
            df.at[numero_linha, campo] = valor

if __name__ == "__main__":
    app = Application()
    app.start_processing()
